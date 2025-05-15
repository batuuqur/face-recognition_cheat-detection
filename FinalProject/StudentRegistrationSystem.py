from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib


lessons_list = ['Math', 'Science', 'History', 'English', 'Programming', 'Art']
original_lessons_list = lessons_list.copy()
selected_lessons = []


background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"


root = Tk()
root.title("Student Registration System")
root.geometry("1250x700+210+100")
root.config(bg=background)

file = pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Semester"
    sheet['D1'] = "Gender"
    sheet['E1'] = "DOB"
    sheet['F1'] = "Date Of Registration"
    sheet['G1'] = "Number"
    sheet['H1'] = "Department"
    sheet['I1'] = "Lessons"
    file.save('Student_data.xlsx')
def Exit():
    root.destroy()


def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title='Select image file', filetype= (("PNG File","*.png"),
                                                                                                        ("JPG File","*.jpg"),
                                                                                                        ("All files","*.txt")))
    img = (Image.open(filename))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2

def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value

    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set("1")

def Clear():
    global img, selected_lessons, lessons_list

    Name.set('')
    DOB.set('')
    Number.set('')
    Department.set('')
    Semester.set("Select Class")
    selected_lessons.clear()
    lesson_entry.config(text='')
    lessons_list = original_lessons_list.copy()
    Class['values'] = lessons_list

    saveButton.config(state='normal')

    img1 =  PhotoImage(file='Resource images/photo.png')
    lbl.config(image=img1)

    lbl.image = img1
    img = ""


def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Semester.get()
    try:
        G1 = gender
    except:
        messagebox.showerror("Error", "Select Gender!")

    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Number.get()
    S1 = Department.get()
    I1 = ', '.join(selected_lessons)

    if N1=="" or C1=="Select Class" or D2=="" or Re1=="" or S1=="" or I1=="":
        messagebox.showerror("error", "Few Data is missing!")
    else:
        file = openpyxl.load_workbook('Student_data.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=C1)
        sheet.cell(column=4, row=sheet.max_row, value=G1)
        sheet.cell(column=5, row=sheet.max_row, value=D2)
        sheet.cell(column=6, row=sheet.max_row, value=D1)
        sheet.cell(column=7, row=sheet.max_row, value=Re1)
        sheet.cell(column=8, row=sheet.max_row, value=S1)
        sheet.cell(column=9, row=sheet.max_row, value=I1)
        file.save(r'Student_data.xlsx')
        try:
            img.save("Student images/"+str(R1)+".png")
        except:
            messagebox.showinfo("Info", "Profile picture is not available!")
        messagebox.showinfo("Info", "Successfully data entered")
        Clear()
        registration_no()

def search():
    global selected_lessons, lessons_list
    text = Search.get()
    Clear()
    saveButton.config(state='disabled')
    #HATA
    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active
    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid", "Invalid registration number!")

    x1 = sheet.cell(row=int(reg_number), column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    Registration.set(x1)
    Name.set(x2)
    Semester.set(x3)
    if x4 == 'Female:':
        R2.select()
    else:
        R1.select()
    DOB.set(x5)
    Date.set(x6)
    Number.set(x7)
    Department.set(x8)
    selected_lessons = x9.split(', ')
    lesson_entry.config(text=', '.join(selected_lessons))
    img = (Image.open("Student images/" + str(x1)+".png"))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image = photo2

    for lesson in selected_lessons:
        if lesson in lessons_list:
            lessons_list.remove(lesson)
    Class['values'] = lessons_list
    Class.set("Select Class")

def Update():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Semester.get()
    selection()
    G1 = gender
    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Number.get()
    S1 = Department.get()
    I1 = ', '.join(selected_lessons)

    if N1=="" or C1=="Select Class" or D2=="" or Re1=="" or S1=="" or I1=="":
        messagebox.showerror("error", "Few Data is missing!")
    else:
        file = openpyxl.load_workbook("Student_data.xlsx")
        sheet = file.active

        for row in sheet.rows:
            if row[0].value == R1:
                name = row[0]
                print(str(name))
                reg_no_position = str(name)[14:-1]
                reg_number = str(name)[15:-1]
                print(reg_number)

        #Don't change this
        #sheet.cell(column=1, row=int(reg_number), value=R1)
        sheet.cell(column=2, row=int(reg_number), value=N1)
        sheet.cell(column=3, row=int(reg_number), value=C1)
        sheet.cell(column=4, row=int(reg_number), value=G1)
        sheet.cell(column=5, row=int(reg_number), value=D2)
        sheet.cell(column=6, row=int(reg_number), value=D1)
        sheet.cell(column=7, row=int(reg_number), value=Re1)
        sheet.cell(column=8, row=int(reg_number), value=S1)
        sheet.cell(column=9, row=int(reg_number), value=I1)

        file.save(r'Student_data.xlsx')

        try:
            img.save("Student images/" + str(R1) + ".png")
        except:
            pass

        messagebox.showinfo("Update", "Update successfully!")
        Clear()


def selection():
    global gender
    value = radio.get()
    if value==1:
        gender="Male"
    else:
        gender="Female"




def get_selected_lessons(event):
    global selected_lessons, lessons_list
    selected_lessons += Class.get().split(', ')
    lesson_entry.config(text=selected_lessons)
    for i in range(len(lessons_list)):
        if lessons_list[i] == Class.get():
            del lessons_list[i]
            break
    # Update the values of the Combobox
    Class['values'] = lessons_list
    Class.set("Select Class")



Label(root, text="Email: batsuqur@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=2, bg="#c36464", fg='#fff', font='arial 20 bold').pack(side=TOP, fill=X)

Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font='arial 20').place(x=820, y=70)
imageicon3 = PhotoImage(file='Resource images/glass.png')
Srch = Button(root, text="Search", compound=LEFT, image=imageicon3, width=123, bg='#68ddfa', font="arial 13 bold", command=search)
Srch.place(x=1060, y=66)

imageicon4 = PhotoImage(file='Resource images/layer.png')
Update_button = Button(root, image=imageicon4, bg="#c36464", command=Update)
Update_button.place(x=110, y=64)

Label(root, text="Registration No:", font="arial 13", fg=framebg, bg=background).place(x=30, y=150)
Label(root, text="Date:", font="arial 13", fg=framebg, bg=background).place(x=500, y=150)

Registration = IntVar()
Date = StringVar()
reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=160, y=150)
registration_no()
today = date.today()
d1 = today.strftime("%d/%m/%y")
date_entry = Entry(root, textvariable=Date, width=15, font='arial 10')
date_entry.place(x=550, y=150)

Date.set(d1)

obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Full Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(obj, text="Date Of Birth:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=100)
Label(obj, text="Gender:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=150)
Label(obj, text="Semester:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(obj, text="Student Number:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=100)
Label(obj, text="Department:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=150)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font="arial 10")
name_entry.place(x=160, y=50)

DOB = StringVar()
dob_entry = Entry(obj, textvariable=DOB, width=20, font="arial 10")
dob_entry.place(x=160, y=100)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=150, y=150)
R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=200, y=150)

Number = StringVar()
religion_entry = Entry(obj, textvariable=Number, width=20, font="arial 10")
religion_entry.place(x=630, y=100)

Department = StringVar()
skill_entry = Entry(obj, textvariable=Department, width=20, font="arial 10")
skill_entry.place(x=630, y=150)

Semester = Combobox(obj, values=['1', '2', '3', '4', '5', '6', '7', '8'], font="Roboto 10", width=17, state="r")
Semester.place(x=630, y=50)
Semester.set("Select Class")

obj2 = LabelFrame(root, text="Lessons to choose", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj2.place(x=30, y=470)

Class = Combobox(obj2, values=lessons_list, state='readonly', width=25)
Class.place(x=30, y=50)
Class.set("Select Class")
Class.bind("<<ComboboxSelected>>", get_selected_lessons)
Class.pack()


obj3 = LabelFrame(root, text="Chosen Lessons", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=150, relief=GROOVE)
obj3.place(x=30, y=530)

lesson_entry = Label(obj3, text='', width=100, font="arial 10", justify=LEFT)
lesson_entry.place(x=30, y=50)


f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

img = PhotoImage(file="Resource images/photo.png")
lbl = Label(f, bg="black", image=img)
lbl.place(x=0, y=0)

Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=showimage).place(x=1000, y=370)
saveButton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=Save)
saveButton.place(x=1000, y=450)
Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=Clear).place(x=1000, y=530)
Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="lightgray", command=Exit).place(x=1000, y=610)
root.mainloop()
