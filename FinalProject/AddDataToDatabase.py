import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
import openpyxl


ref = db.reference("Students")

workbook = openpyxl.load_workbook('Student_data.xlsx')
worksheet = workbook['Sheet']
last_row = worksheet.max_row

data = {}

for row_number in range(2, last_row + 1):
    id = str(worksheet.cell(row=row_number, column=1).value)
    student_data = {
        "Name": worksheet.cell(row=row_number, column=2).value,
        "Semester": worksheet.cell(row=row_number, column=3).value,
        "Gender": worksheet.cell(row=row_number, column=4).value,
        "DOB": worksheet.cell(row=row_number, column=5).value,
        "Date Of Registration": worksheet.cell(row=row_number, column=6).value,
        "Number": worksheet.cell(row=row_number, column=7).value,
        "Department": worksheet.cell(row=row_number, column=8).value,
        "Lessons": worksheet.cell(row=row_number, column=9).value,
        "last_attendance_time": '2023-04-08 18:10:28'
    }
    data[id] = student_data

print(data)

for key, value in data.items():
    ref.child(key).set(value)